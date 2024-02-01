import kivy
from kivy import platform
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from datetime import timedelta, date
from kivy.clock import Clock
from kivymd.app import MDApp
from kivymd.uix.floatlayout import FloatLayout
from kivymd.uix.screen import MDScreen
#from kivy.uix.image import Image
from kivymd.uix.button import MDRectangleFlatButton, MDRaisedButton
from kivymd.uix.button import MDIconButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.selectioncontrol import MDCheckbox

class MDScreen(FloatLayout):

    cb1_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
    cb2_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
    cb3_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
    cb4_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
    
    cb1_pre_pct_list = []
    cb1_chk_pct_list = []
    cb1_iti_pct_list = []
    cb1_shk_pct_list = []

    cb2_pre_pct_list = []
    cb2_chk_pct_list = []
    cb2_iti_pct_list = []
    cb2_shk_pct_list = []

    cb3_pre_pct_list = []
    cb3_chk_pct_list = []
    cb3_iti_pct_list = []
    cb3_shk_pct_list = []

    cb4_pre_pct_list = []
    cb4_chk_pct_list = []
    cb4_iti_pct_list = []
    cb4_shk_pct_list = []

    list_index = 0
    watch_started = False # whether the  timer is started or not
    datbtn = 0
    datbtnfasheet = -2
    e = 2
    time_var = 0
    
    def set_default_lists(self):
        self.cb1_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
        self.cb2_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
        self.cb3_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
        self.cb4_list = [0,0,0,0,0,0, 0,0,0,0,0,0, 0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0,
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 
                    0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0, 0,0,0,0,0,0,0]
        
        self.cb1_pre_pct_list = []
        self.cb1_chk_pct_list = []
        self.cb1_iti_pct_list = []
        self.cb1_shk_pct_list = []

        self.cb2_pre_pct_list = []
        self.cb2_chk_pct_list = []
        self.cb2_iti_pct_list = []
        self.cb2_shk_pct_list = []

        self.cb3_pre_pct_list = []
        self.cb3_chk_pct_list = []
        self.cb3_iti_pct_list = []
        self.cb3_shk_pct_list = []

        self.cb4_pre_pct_list = []
        self.cb4_chk_pct_list = []
        self.cb4_iti_pct_list = []
        self.cb4_shk_pct_list = []

        self.list_index = 0

        self.time_var = 0

    # create main GUI screen
    def __init__(self):
        super(MDScreen, self).__init__()

    def checkdaboxes(self):
        
        self.cb1_list[self.list_index] = 1 if self.ids['cb1'].active else 0

        self.cb2_list[self.list_index] = 1 if self.ids['cb2'].active else 0

        self.cb3_list[self.list_index] = 1 if self.ids['cb3'].active else 0

        self.cb4_list[self.list_index] = 1 if self.ids['cb4'].active else 0

        self.list_index += 1

        self.time_var += 10

        self.check_text.text = str('Next\nCheck\n@')

        minutes = self.time_var // 60
        seconds = self.time_var % 60
        self.time_guide.text = str(f'{minutes}:{seconds:02d}')
    
        print(f'{minutes}:{seconds:02d}')

        print(self.list_index)

        self.ids['cb1'].active = False
        self.ids['cb2'].active = False
        self.ids['cb3'].active = False
        self.ids['cb4'].active = False
        
    def calc_pct(self):
        chunk_size = 6
        post_chk = 7
        post_start_index = 19
        shk_index = 18

        for i in range(3): # Quantifiy pct FZ for first 3min
            start_index = i * chunk_size
            end_index = start_index + chunk_size
            cb1_sum_min = sum(self.cb1_list[start_index:end_index])
            cb1_min_pct = round((cb1_sum_min / 6) * 100)
            self.cb1_pre_pct_list.append(cb1_min_pct)

            cb2_sum_min = sum(self.cb2_list[start_index:end_index])
            cb2_min_pct = round((cb2_sum_min / 6) * 100)
            self.cb2_pre_pct_list.append(cb2_min_pct)

            cb3_sum_min = sum(self.cb3_list[start_index:end_index])
            cb3_min_pct = round((cb3_sum_min / 6) * 100)
            self.cb3_pre_pct_list.append(cb3_min_pct)

            cb4_sum_min = sum(self.cb4_list[start_index:end_index])
            cb4_min_pct = round((cb4_sum_min / 6) * 100)
            self.cb4_pre_pct_list.append(cb4_min_pct)

        for i in range(25): 
            # Quantifiy pct FZ for inter trial intervals and shock intervals combined 25min
            post_end_index = post_start_index + post_chk

            cb1_post_sum_chk = sum(self.cb1_list[post_start_index:post_end_index])
            cb1_chk_pct = round((cb1_post_sum_chk / 7) * 100)
            self.cb1_chk_pct_list.append(cb1_chk_pct)

            cb2_post_sum_chk = sum(self.cb2_list[post_start_index:post_end_index])
            cb2_chk_pct = round((cb2_post_sum_chk / 7) * 100)
            self.cb2_chk_pct_list.append(cb2_chk_pct)

            cb3_post_sum_chk = sum(self.cb3_list[post_start_index:post_end_index])
            cb3_chk_pct = round((cb3_post_sum_chk / 7) * 100)
            self.cb3_chk_pct_list.append(cb3_chk_pct)

            cb4_post_sum_chk = sum(self.cb4_list[post_start_index:post_end_index])
            cb4_chk_pct = round((cb4_post_sum_chk / 7) * 100)
            self.cb4_chk_pct_list.append(cb4_chk_pct)

            # Quantifiy pct FZ for inter trial intervals 25min
            post_end_index = post_start_index + chunk_size
            cb1_post_sum_min = sum(self.cb1_list[post_start_index:post_end_index])
            cb1_min_pct = round((cb1_post_sum_min / 6) * 100)
            self.cb1_iti_pct_list.append(cb1_min_pct)

            cb2_post_sum_min = sum(self.cb2_list[post_start_index:post_end_index])
            cb2_min_pct = round((cb2_post_sum_min / 6) * 100)
            self.cb2_iti_pct_list.append(cb2_min_pct)

            cb3_post_sum_min = sum(self.cb3_list[post_start_index:post_end_index])
            cb3_min_pct = round((cb3_post_sum_min / 6) * 100)
            self.cb3_iti_pct_list.append(cb3_min_pct)

            cb4_post_sum_min = sum(self.cb4_list[post_start_index:post_end_index])
            cb4_min_pct = round((cb4_post_sum_min / 6) * 100)
            self.cb4_iti_pct_list.append(cb4_min_pct)

            # Quantify pct FZ for shock intervals 25 X 10sec shocks
            cb1_shk_val = self.cb1_list[shk_index]
            cb1_shk_pct = cb1_shk_val*100
            self.cb1_shk_pct_list.append(cb1_shk_pct)

            cb2_shk_val = self.cb2_list[shk_index]
            cb2_shk_pct = cb2_shk_val*100
            self.cb2_shk_pct_list.append(cb2_shk_pct)

            cb3_shk_val = self.cb3_list[shk_index]
            cb3_shk_pct = cb3_shk_val*100
            self.cb3_shk_pct_list.append(cb3_shk_pct)

            cb4_shk_val = self.cb4_list[shk_index]
            cb4_shk_pct = cb4_shk_val*100
            self.cb4_shk_pct_list.append(cb4_shk_pct)

            shk_index += 7
            post_start_index += 7

    # re-create main GUI screen to reset app inputs
    def reset_func(self):
        
        FloatLayout.clear_widgets(self)
        
        super(MDScreen, self).__init__()
        
    # create/save the Excel workbook
    def mk_xl_file(self):
        try:
            file_name = self.ids.file_name.text
            wb = Workbook()
            sheet1 = wb.create_sheet('FZnDATA',0)
            
            sheet1['A1'] = 'File'
            sheet1['B1'] = 'Obsvr.'
            sheet1['C1'] = 'Exp.'
            sheet1['D1'] = 'Date'
            sheet1['E1'] = 'Subject'
            sheet1['F1'] = 'Pre 1'
            sheet1['G1'] = 'Pre 2'
            sheet1['H1'] = 'Pre 3'
            sheet1['I1'] = 'ITI+CS 1'
            sheet1['J1'] = 'ITI+CS 2'
            sheet1['K1'] = 'ITI+CS 3'
            sheet1['L1'] = 'ITI+CS 4'
            sheet1['M1'] = 'ITI+CS 5'
            sheet1['N1'] = 'ITI+CS 6'
            sheet1['O1'] = 'ITI+CS 7'
            sheet1['P1'] = 'ITI+CS 8'
            sheet1['Q1'] = 'ITI+CS 9'
            sheet1['R1'] = 'ITI+CS 10'
            sheet1['S1'] = 'ITI+CS 11'
            sheet1['T1'] = 'ITI+CS 12'
            sheet1['U1'] = 'ITI+CS 13'
            sheet1['V1'] = 'ITI+CS 14'
            sheet1['W1'] = 'ITI+CS 15'
            sheet1['X1'] = 'ITI+CS 16'
            sheet1['Y1'] = 'ITI+CS 17'
            sheet1['Z1'] = 'ITI+CS 18'
            sheet1['AA1'] = 'ITI+CS 19'
            sheet1['AB1'] = 'ITI+CS 20'
            sheet1['AC1'] = 'ITI+CS 21'
            sheet1['AD1'] = 'ITI+CS 22'
            sheet1['AE1'] = 'ITI+CS 23'
            sheet1['AF1'] = 'ITI+CS 24'
            sheet1['AG1'] = 'ITI+CS 25'
            sheet1['AH1'] = 'ITI 1'
            sheet1['AI1'] = 'ITI 2'
            sheet1['AJ1'] = 'ITI 3'
            sheet1['AK1'] = 'ITI 4'
            sheet1['AL1'] = 'ITI 5'
            sheet1['AM1'] = 'ITI 6'
            sheet1['AN1'] = 'ITI 7'
            sheet1['AO1'] = 'ITI 8'
            sheet1['AP1'] = 'ITI 9'
            sheet1['AQ1'] = 'ITI 10'
            sheet1['AR1'] = 'ITI 11'
            sheet1['AS1'] = 'ITI 12'
            sheet1['AT1'] = 'ITI 13'
            sheet1['AU1'] = 'ITI 14'
            sheet1['AV1'] = 'ITI 15'
            sheet1['AW1'] = 'ITI 16'
            sheet1['AX1'] = 'ITI 17'
            sheet1['AY1'] = 'ITI 18'
            sheet1['AZ1'] = 'ITI 19'
            sheet1['BA1'] = 'ITI 20'
            sheet1['BB1'] = 'ITI 21'
            sheet1['BC1'] = 'ITI 22'
            sheet1['BD1'] = 'ITI 23'
            sheet1['BE1'] = 'ITI 24'
            sheet1['BF1'] = 'ITI 25'
            sheet1['BG1'] = 'CS 1'
            sheet1['BH1'] = 'CS 2'
            sheet1['BI1'] = 'CS 3'
            sheet1['BJ1'] = 'CS 4'
            sheet1['BK1'] = 'CS 5'
            sheet1['BL1'] = 'CS 6'
            sheet1['BM1'] = 'CS 7'
            sheet1['BN1'] = 'CS 8'
            sheet1['BO1'] = 'CS 9'
            sheet1['BP1'] = 'CS 10'
            sheet1['BQ1'] = 'CS 11'
            sheet1['BR1'] = 'CS 12'
            sheet1['BS1'] = 'CS 13'
            sheet1['BT1'] = 'CS 14'
            sheet1['BU1'] = 'CS 15'
            sheet1['BV1'] = 'CS 16'
            sheet1['BW1'] = 'CS 17'
            sheet1['BX1'] = 'CS 18'
            sheet1['BY1'] = 'CS 19'
            sheet1['BZ1'] = 'CS 20'
            sheet1['CA1'] = 'CS 21'
            sheet1['CB1'] = 'CS 22'
            sheet1['CC1'] = 'CS 23'
            sheet1['CD1'] = 'CS 24'
            sheet1['CE1'] = 'CS 25'
        
            wb.save('f:'+file_name+'.xlsx')

            self.file_name_err.text = str('')
            self.ids['xlfile_btn'].disabled = True
            self.ids['unlock_data'].disabled = False
            if self.ids.file_name.text == str(''):
                self.file_name_err.text = str("File name field cannot be empty")
                self.ids['xlfile_btn'].disabled = False
                self.ids['unlock_data'].disabled = True
            # Check for invalid characters
            invalid_chars = "/\?*[]"
            for char in invalid_chars:
                if char in file_name:
                    self.file_name_err.text = str("/ \ ? * [ ] are invalid filename chars")
                    self.ids['xlfile_btn'].disabled = False
                    self.ids['unlock_data'].disabled = True
            # Check for invalid starting characters
            invalid_start_chars = "~!"
            if file_name[0] == "." or file_name[0] in invalid_start_chars:
                self.file_name_err.text = str("~ ! are invalid starting filename chars")
                self.ids['xlfile_btn'].disabled = False
                self.ids['unlock_data'].disabled = True   
        except ValueError: 
            self.file_name_err.text = str("Invalid Excel Filename")
            self.ids['xlfile_btn'].disabled = False
            self.ids['unlock_data'].disabled = True

    def unlock_unlock_btn(self):
        self.ids['unlock_btn'].disabled = False

    def unlock_submit_btn(self):
        self.ids['submit_dat_btn'].disabled = False

    # create and write data to excel file    
    def submit_data(self):
        
        today = date.today()

        file_name = self.ids.file_name.text
        wb = load_workbook('f:'+file_name+'.xlsx')
        #wb = load_workbook('//storage//emulated//0//Download//'+file_name+'.xlsx')
        sheet1 = wb['FZnDATA']
        sheet1['A'+str(self.e)] = str(self.ids.file_name.text)
        sheet1['B'+str(self.e)] = str(self.ids.scorerID.text)
        sheet1['C'+str(self.e)] = str(self.ids.experimentID.text)
        sheet1['D'+str(self.e)] = str(today)
        sheet1['E'+str(self.e)] = str(self.ids.cb1_subID.text)
        sheet1['E'+str(self.e+1)] = str(self.ids.cb2_subID.text)
        sheet1['E'+str(self.e+2)] = str(self.ids.cb3_subID.text)
        sheet1['E'+str(self.e+3)] = str(self.ids.cb4_subID.text)



        for i, value in enumerate(self.cb1_pre_pct_list, start=6):
            sheet1.cell(row=self.e, column=i, value=(value))
        for i, value in enumerate(self.cb1_chk_pct_list, start=9):
            sheet1.cell(row=self.e, column=i, value=(value))
        for i, value in enumerate(self.cb1_iti_pct_list, start=34):
            sheet1.cell(row=self.e, column=i, value=(value))
        for i, value in enumerate(self.cb1_shk_pct_list, start=59):
            sheet1.cell(row=self.e, column=i, value=(value))

        for i, value in enumerate(self.cb2_pre_pct_list, start=6):
            sheet1.cell(row=self.e+1, column=i, value=(value))
        for i, value in enumerate(self.cb2_chk_pct_list, start=9):
            sheet1.cell(row=self.e+1, column=i, value=(value))
        for i, value in enumerate(self.cb2_iti_pct_list, start=34):
            sheet1.cell(row=self.e+1, column=i, value=(value))
        for i, value in enumerate(self.cb2_shk_pct_list, start=59):
            sheet1.cell(row=self.e+1, column=i, value=(value))
        
        for i, value in enumerate(self.cb3_pre_pct_list, start=6):
            sheet1.cell(row=self.e+2, column=i, value=(value))
        for i, value in enumerate(self.cb3_chk_pct_list, start=9):
            sheet1.cell(row=self.e+2, column=i, value=(value))
        for i, value in enumerate(self.cb3_iti_pct_list, start=34):
            sheet1.cell(row=self.e+2, column=i, value=(value))
        for i, value in enumerate(self.cb3_shk_pct_list, start=59):
            sheet1.cell(row=self.e+2, column=i, value=(value))
        
        for i, value in enumerate(self.cb4_pre_pct_list, start=6):
            sheet1.cell(row=self.e+3, column=i, value=(value))
        for i, value in enumerate(self.cb4_chk_pct_list, start=9):
            sheet1.cell(row=self.e+3, column=i, value=(value))
        for i, value in enumerate(self.cb4_iti_pct_list, start=34):
            sheet1.cell(row=self.e+3, column=i, value=(value))
        for i, value in enumerate(self.cb4_shk_pct_list, start=59):
            sheet1.cell(row=self.e+3, column=i, value=(value))
    
        self.e+=4

        # save the xcel workbook
        wb.save('f:'+file_name+'.xlsx')
        wb.close()

        # reset all values on display screen
        self.ids['submit_dat_btn'].disabled = True
        self.ids['unlock_data'].disabled = False
        self.ids['cb1'].active = False
        self.ids['cb2'].active = False
        self.ids['cb3'].active = False
        self.ids['cb4'].active = False

        self.set_default_lists()

    # func to save user defined session time
    def set_sess_time(self):
        self.seconds_set = str(self.ids.usr_sec.text)
        self.minutes_set = str(self.ids.usr_min.text)
        # set time only if time is not zero
        if self.seconds_set == 0 and self.minutes_set == 0 or self.seconds_set == '00' and self.minutes_set == '00':
            self.ids['play_pause_btn'].disabled = True
            self.ids['reset_btn'].disabled = True
        else:
           self.ids['play_pause_btn'].disabled = False
 
    # func to set user defined session time
    def get_usr_min_sec(self):
        seconds = str(self.ids.usr_sec.text)
        minutes = str(self.ids.usr_min.text)
        # save time only if time is not zero
        if self.seconds_set == 0 and self.minutes_set == 0 or self.seconds_set == '00' and self.minutes_set == '00':
            self.ids['play_pause_btn'].disabled = True
        else:
            self.usr_sec_in = seconds
            self.usr_min_in = minutes
    
    # func to increment by seconds and time to str
    def get_string_time(self, dt):
        self.increment_seconds()
        self.ids['bout_btn'].disabled = False

        minutes = str(self.minutes)
        seconds = str(self.seconds)
        #print(seconds, minutes)

        if len(seconds) < 2:
            seconds = '0' + seconds

        if len(minutes) < 2:
            minutes = '0' + minutes

        self.usr_min.text = str(minutes)
        self.usr_sec.text = str(seconds)

        if self.seconds == 0 and self.minutes == 0:
            self.ids['cb1'].disabled = True
            self.ids['cb2'].disabled = True
            self.ids['cb3'].disabled = True
            self.ids['cb4'].disabled = True
            self.ids['play_pause_btn'].icon = 'play'
            self.ids['reset_btn'].disabled = False

    def increment_seconds(self):
        """Increment the seconds by 1 second"""
        self.seconds = int(self.ids.usr_sec.text)
        self.minutes = int(self.ids.usr_min.text)

        if self.seconds == 0 and self.minutes != 0:
            self.increment_minutes()
            self.seconds = 59
        elif self.seconds == 0 and self.minutes == 0: 
            Clock.unschedule(self.get_string_time)
        else:
            self.seconds -= 1
            
    def increment_minutes(self):
        """Increment the minutes by 1 minute"""
        self.minutes = int(self.ids.usr_min.text)
        self.minutes -= 1

    # function to start stopwatch if not running otherwise stop it
    def start_or_stop_stopwatch(self):
        if self.watch_started:
            self.watch_started = False
            self.ids['play_pause_btn'].icon = 'play'
            self.ids['reset_btn'].disabled = False
            self.ids['cb1'].disabled = True
            self.ids['cb2'].disabled = True
            self.ids['cb3'].disabled = True
            self.ids['cb4'].disabled = True
            # stop timer
            Clock.unschedule(self.get_string_time)
        else:
            self.watch_started = True
            self.ids['play_pause_btn'].icon = 'pause'
            self.ids['reset_btn'].disabled = True
            self.ids['sess_tm_btn'].disabled = True
            # start timer / schedule to update get_string_time func every 1s
            Clock.schedule_interval(self.get_string_time, 1)

    # func to reset stopwatch and variables
    def reset_stopwatch(self):
        """Set the stopwatch to user defined session time"""
        self.usr_min.text = str(self.minutes_set)
        self.usr_sec.text = str(self.seconds_set)
        # reset all values on display screen
        self.bouts = []
        self.times = []
        self.en_times = []
        self.press_count = 0
        self.sum_times = 0
        self.ids['reset_btn'].disabled = True
        self.ids['cb1'].disabled = True
        self.ids['cb2'].disabled = True
        self.ids['cb3'].disabled = True
        self.ids['cb4'].disabled = True
        self.ids['sess_tm_btn'].disabled = False

class FRZNTRACK(MDApp):
    
    # build app
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Pink"

        #app icon image
        self.icon = 'ratice.png'

        return MDScreen()

FRZNTRACK = FRZNTRACK()
FRZNTRACK.run()